import os
import streamlit as st
import pandas as pd
import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from demo_data import get_demo_test_cases

from generator import generate_test_cases
from history import save_to_history, load_history, load_from_history

DEMO_MODE = os.environ.get("QA_COPILOT_DEMO", "0").lower() in ("1", "true", "yes")

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="QA Copilot",
    page_icon="🤖",
    layout="wide"
)

# ── Styles ───────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main-title { font-size: 2.2rem; font-weight: 800; color: #1F3864; }
    .sub-title  { font-size: 1rem; color: #555; margin-bottom: 1.5rem; }
    .tc-card    { background: #f8f9fa; border-radius: 10px;
                  padding: 1rem; margin-bottom: 1rem;
                  border-left: 5px solid #1F3864; }
    .badge-high   { background:#FFCCCC; color:#c0392b;
                    padding:2px 10px; border-radius:12px; font-size:0.8rem; font-weight:600; }
    .badge-medium { background:#FFF2CC; color:#d68910;
                    padding:2px 10px; border-radius:12px; font-size:0.8rem; font-weight:600; }
    .badge-low    { background:#CCFFCC; color:#1e8449;
                    padding:2px 10px; border-radius:12px; font-size:0.8rem; font-weight:600; }
</style>
""", unsafe_allow_html=True)

# ── Helpers ───────────────────────────────────────────────────────────────────
PRIORITY_COLORS = {"High": "FFCCCC", "Medium": "FFF2CC", "Low": "CCFFCC"}
HEADER_COLOR    = "1F3864"

def format_step(i, text):
    clean = re.sub(r'^[Ss]tep\s*\d+\s*[:\-]\s*', '', text).strip()
    return f"{i+1} - {clean}"

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def build_csv(test_cases):
    rows = []
    for tc in test_cases:
        for i, step in enumerate(tc.steps):
            paso = format_step(i, step)
            if i == 0:
                rows.append([tc.tc_id, tc.title, tc.tc_type, "Manual",
                              tc.priority, tc.preconditions, paso, tc.expected_result])
            else:
                rows.append(["", "", "", "", "", "", paso, ""])
        rows.append([""] * 8)
    cols = ["ID CP","Nombre CP","Descripcion","Tipo de Prueba",
            "Prioridad","Precondiciones","Paso","Resultado Esperado"]
    return pd.DataFrame(rows, columns=cols).to_csv(index=False).encode("utf-8")

def build_xlsx(test_cases, req_id):
    wb = Workbook()
    ws = wb.active
    ws.title = req_id[:30]
    headers = ["ID CP","Nombre CP","Descripcion","Tipo de Prueba",
               "Prioridad","Precondiciones","Paso","Resultado Esperado"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", start_color=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    row = 2
    for tc in test_cases:
        bg = PRIORITY_COLORS.get(tc.priority, "FFFFFF")
        start_row = row
        for i, step in enumerate(tc.steps):
            vals = ([tc.tc_id, tc.title, tc.tc_type, "Manual",
                     tc.priority, tc.preconditions, format_step(i, step), tc.expected_result]
                    if i == 0 else ["","","","","","", format_step(i, step),""])
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font      = Font(name="Arial", size=10)
                cell.fill      = PatternFill("solid", start_color=bg)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border    = thin_border()
            ws.row_dimensions[row].height = 40
            row += 1
        end_row = row - 1
        if end_row > start_row:
            for col in list(range(1, 7)) + [8]:
                ws.merge_cells(start_row=start_row, start_column=col,
                               end_row=end_row, end_column=col)
                c = ws.cell(row=start_row, column=col)
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for col, w in enumerate([10,30,15,14,12,40,55,50], 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def badge(priority):
    cls = f"badge-{priority.lower()}"
    return f'<span class="{cls}">{priority}</span>'

# ── Sidebar — History ─────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 📋 History")
    history = load_history()
    if not history:
        st.caption("No history yet.")
    else:
        for req_id, data in history.items():
            tc_count = len(data["test_cases"])
            if st.button(f"📄 {req_id} ({tc_count} TCs)", key=req_id, use_container_width=True):
                result = load_from_history(req_id)
                if result:
                    st.session_state.test_cases  = result[0]
                    st.session_state.current_req = req_id
                    st.rerun()

# ── Main ──────────────────────────────────────────────────────────────────────
st.markdown('<p class="main-title">🤖 QA Copilot</p>', unsafe_allow_html=True)
st.markdown('<p class="sub-title">AI-powered enhanced test case generator</p>',
            unsafe_allow_html=True)

col1, col2 = st.columns([1, 2])

with col1:
    req_id     = st.text_input("Requirement ID", placeholder="e.g. FB-00312")
    user_story = st.text_area("User Story", height=220,
                              placeholder="As a user, I want to...")
    use_demo   = st.checkbox(
        "🔧 Use demo mode (no Gemini API key required)",
        value=DEMO_MODE or not os.environ.get("GOOGLE_API_KEY")
    )
    generate   = st.button("⚡ Generate Test Cases", use_container_width=True, type="primary")

if generate:
    if not req_id or not user_story:
        st.warning("Please fill in both fields.")
    else:
        if use_demo:
            st.warning("⚠️ Running in demo mode: Gemini is not used.")

        with st.spinner("Generating test cases, please be patient..."):
            try:
                tcs = generate_test_cases(req_id, user_story, demo_mode=use_demo)

            except Exception as e:

                if (
                    "Insufficient Balance" in str(e)
                    or
                    "API key" in str(e)
                ):

                    st.warning(
                        "⚠️ AI provider unavailable. Running in Demo Mode."
                    )

                    tcs = get_demo_test_cases(user_story)

                else:
                    st.error(f"Error: {e}")
                    st.stop()

            st.session_state.test_cases = tcs
            st.session_state.current_req = req_id
            save_to_history(req_id, user_story, tcs)
            st.rerun()

# ── Results ───────────────────────────────────────────────────────────────────
if "test_cases" in st.session_state and st.session_state.test_cases:
    tcs     = st.session_state.test_cases
    cur_req = st.session_state.current_req

    with col2:
        st.markdown(f"### ✅ {len(tcs)} Test Cases — `{cur_req}`")

        dl1, dl2 = st.columns(2)
        with dl1:
            st.download_button("⬇️ Download CSV", build_csv(tcs),
                               file_name=f"{cur_req}_test_cases.csv",
                               mime="text/csv", use_container_width=True)
        with dl2:
            st.download_button("⬇️ Download Excel", build_xlsx(tcs, cur_req),
                               file_name=f"{cur_req}_test_cases.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)

        st.divider()

        for tc in tcs:
            with st.expander(f"**{tc.tc_id}** — {tc.title}"):
                c1, c2, c3 = st.columns(3)
                c1.markdown(f"**Type:** {tc.tc_type}")
                c2.markdown(f"**Priority:** {badge(tc.priority)}", unsafe_allow_html=True)
                c3.markdown(f"**Steps:** {len(tc.steps)}")
                st.markdown(f"**Preconditions:** {tc.preconditions}")
                st.markdown("**Steps:**")
                for i, step in enumerate(tc.steps):
                    st.markdown(f"- {format_step(i, step)}")
                st.success(f"✅ {tc.expected_result}")
